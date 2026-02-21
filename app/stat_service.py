import datetime
from openpyxl import Workbook
from fastapi.responses import StreamingResponse
from openpyxl.utils import get_column_letter
from io import BytesIO

from .classes import ApiConfig, StatSchema
from .binance import get_balance
from .database import get_yesterday_balance_db, get_stat_history_db
from .config import bots_list
from .logger import get_logger


logger = get_logger(__name__)


async def make_stat_file(bot_id: str) -> StreamingResponse:
    stat_data = await get_stat_history_db(bot_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "PnL history"

    ws.append(["Дата", "Баланс", "PNL", "PNL %"])

    for item in stat_data:
        ws.append([
            item.date,                 # datetime.date
            float(item.balance),       # number
            float(item.pnl),           # number
            float(item.pnl_percent),   # number
        ])

    # formats: A date, B/C money, D percent-value (как у тебя, не *100 и не /100)
    col_formats = {"A": "DD.MM.YYYY", "B": "0.00", "C": "0.00", "D": "0.00"}
    for col, fmt in col_formats.items():
        for cell in ws[col][1:]:
            cell.number_format = fmt

    for i, w in enumerate((14, 14, 12, 10), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="stat_history_{bot_id}.xlsx"'},
    )


async def count_day_stat(bot_id: str):
    bot = next(b for b in bots_list.bots if b.key == bot_id)

    api = ApiConfig(
        key=bot.api.key,
        secret=bot.api.secret,
    )

    balance = get_balance(api)
    yesterday = datetime.date.today() - datetime.timedelta(days=1)

    yesterday_balance = await get_yesterday_balance_db(bot_id, yesterday)

    pnl_value = balance - yesterday_balance
    pnl_percent = (pnl_value / 800 * 100)

    return StatSchema(
        balance=balance,
        pnl=pnl_value,
        pnl_percent=pnl_percent,
    )


async def count_all_balance():
    yesterday = datetime.date.today() - datetime.timedelta(days=1)

    total_balance = 0
    total_yesterday_balance = 0

    bots = bots_list.bots

    for bot in bots:
        api = ApiConfig(
            key=bot.api.key,
            secret=bot.api.secret,
        )

        balance = get_balance(api)
        total_balance += float(balance)

        yesterday_balance = await get_yesterday_balance_db(bot.key, yesterday)

        total_yesterday_balance += float(yesterday_balance)

    pnl_value = total_balance - total_yesterday_balance

    base = 800 * len(bots)
    pnl_percent = (pnl_value / base * 100)

    return StatSchema(
        balance=float(total_balance),
        pnl=float(pnl_value),
        pnl_percent=float(pnl_percent),
    )